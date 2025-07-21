from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from knowledge_series.models import BookWishes
from dr_gift_catalogs.models import DrGiftCatalog
from django.db.models import Q
from django.core.paginator import Paginator
import openpyxl , os, zipfile, shutil, json
from io import BytesIO
from django.http import HttpResponse
from core.models import Territory, UserProfile
from django.conf import settings
from anniversary.models import Anniversary
from green_corner.models import GreenCorner
from .models import AccessControl
from doctors_opinion.models import DoctorOpinion
from openpyxl.styles import Alignment
from doctors_data.models import Doctor, Chamber
from . import utils

# Create your views here.
@login_required
def index(request):
    return render(request, 'index.html')

@login_required
def knowledge_series(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        page_number = int(request.GET.get('page') or 1)
        per_page = int(request.GET.get("per_page") or 10)
        sort = request.GET.get("sort", "territory")
        direction = request.GET.get("direction", "asc")        
        # Get data usnig utils filter function
        data = utils.filter_knowledge_series_data(request)
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)
    return render(request, 'knowledge_series.html', {
        'data': page_obj, 'search_query': search_query, 'per_page': per_page, 'sort': sort, 'direction': direction
    })
    
@login_required 
def export_knowledge_series(request):
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Knowledge Series Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Book']
    worksheet.append(headers)
    
    # Get data using the utils function
    data = utils.filter_knowledge_series_data(request)
    
    # Populate the worksheet with data
    for obj in data:
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            obj.book
        ]
        worksheet.append(row)
    
    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create a response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="knowledge_series_data.xlsx"'
    
    return response

@login_required
def delete_knowledge_series_data(request, id):
    territory = request.user.username
    if request.user.is_superuser:
        territory = request.GET.get('territory')
    # territory_obj = Territory.objects.get(territory=territory)
    # ks_obj = BookWishes.objects.filter(territory__territory=territory)
    obj = BookWishes.objects.get(id=id)
    try:
        obj.delete()
        
        messages.success(request, f"Data deleted successfully.")
        if request.user.is_superuser:
            return redirect('knowledge_series')
        return redirect('home')
    except BookWishes.DoesNotExist:
        messages.error(request, "Invalid Item selected.")
        if request.user.is_superuser:
            return redirect('knowledge_series')
        return redirect('home')
    
@login_required 
def download_gift_catalogs(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Gift Catalogs Data"
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Gift Choice']
    worksheet.append(headers)
    data = utils.filter_gift_catalogs_data(request)
    image_paths = set()
    for obj in data:
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            obj.gift
        ]
        worksheet.append(row)
        if obj.conference_image:
            image_path = os.path.join(settings.MEDIA_ROOT, obj.conference_image.name)
            if os.path.exists(image_path):
                image_paths.add(image_path)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for image_path in image_paths:
            relative_path = os.path.relpath(image_path, settings.MEDIA_ROOT)
            zip_file.write(image_path, relative_path)
        # Add the Excel file
        zip_file.writestr('gift_catalogs_data.xlsx', buffer.getvalue())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="gift_catalogs_data_bundle.zip"'
    return response

@login_required 
def export_gift_catalogs(request):
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Gift Catalogs Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Gift Choice']
    worksheet.append(headers)
    data = utils.filter_gift_catalogs_data(request)
    # Populate the worksheet with data
    for obj in data:
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            obj.gift
        ]
        worksheet.append(row)
    
    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create a response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="gift_catalogs_data.xlsx"'
    
    return response

@login_required
def gift_catalogs(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        page_number = int(request.GET.get('page') or 1)
        per_page = int(request.GET.get("per_page") or 10)
        sort = request.GET.get("sort", "territory")
        direction = request.GET.get("direction", "asc")
        # Get data using the utils function
        data = utils.filter_gift_catalogs_data(request)
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)    
    return render(request, 'gift_catalogs.html',{'data':page_obj, 'search_query':search_query, 'per_page':per_page, 'sort':sort, 'direction':direction})  

@login_required
def anniversary(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        page_number = int(request.GET.get('page') or 1)
        per_page = int(request.GET.get("per_page") or 10)
        sort = request.GET.get("sort", "territory")
        direction = request.GET.get("direction", "asc")
        # Get data using the utils function
        data = utils.filter_anniversary_data(request)
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)    
    return render(request, 'anniversary.html',{'data':page_obj, 'search_query':search_query, 'per_page':per_page, 'sort':sort, 'direction':direction}) 

@login_required 
def export_anniversary(request):
    """
    Export the Anniversary(Enlighted Together) data to an Excel file.
    """
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Anniversary Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Anniversary Date']
    worksheet.append(headers)
    
    # Get data using the utils function
    data = utils.filter_anniversary_data(request)
    
    # Populate the worksheet with data
    for obj in data:
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            obj.anniversary_date.strftime('%Y-%m-%d')
        ]
        worksheet.append(row)
    
    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create a response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="anniversary_data.xlsx"'
    
    return response

@login_required 
def download_anniverysary(request):
    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Anniversary Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Anniversary Date']
    worksheet.append(headers)
    # Get data using the utils function
    data = utils.filter_anniversary_data(request)
    # Collect image paths for filtering
    image_paths = set()
    # Populate the worksheet with data
    for obj in data:
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            obj.anniversary_date
        ]
        worksheet.append(row)
        if obj.image:  # Check if image exists
            image_path = os.path.join(settings.MEDIA_ROOT, obj.image.name)
            if os.path.exists(image_path):
                image_paths.add(image_path)
                
    # Save the workbook to a BytesIO (in memory) object
    buffer =BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create zip in memory
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Add filteres images to the zip file
        for image_path in image_paths:
            relative_path = os.path.relpath(image_path, settings.MEDIA_ROOT)
            zip_file.write(image_path, relative_path)    
        # Add the Excel file
        zip_file.writestr('anniversary_data.xlsx', buffer.getvalue())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="anniversary_data_bundle.zip"'
    return response

@login_required
def green_corner(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        page_number = int(request.GET.get('page') or 1)
        per_page = int(request.GET.get("per_page") or 10)
        sort = request.GET.get("sort", "territory")
        direction = request.GET.get("direction", "asc")        
        data = utils.filter_green_corner_data(request)
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)
    return render(request, 'green_corner.html', {
        'data': page_obj, 'search_query': search_query, 'per_page': per_page, 'sort': sort, 'direction': direction
    })

@login_required 
def export_rgc(request):
    """
    Export the knowledge series data to an Excel file.
    """

    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Radiant Green Corner Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Flower Plants', 'Medicinal Plants']
    worksheet.append(headers)
    
    # Populate the worksheet with data
    queryset = GreenCorner.objects.select_related('territory')
    try:
        profile = request.user.userprofile
        if profile.user_type == 'zone':
            queryset = queryset.filter(territory__zone_name=profile.zone_name)
        elif  profile.user_type == 'region':
            queryset = queryset.filter(territory__region_name=profile.region_name)
    except UserProfile.DoesNotExist:
        if not request.user.is_superuser:
            queryset = GreenCorner.objects.none()
    for obj in queryset:
        flower_plants = f'{obj.first_flower_plant}, {obj.second_flower_plant}, {obj.third_flower_plant}.'
        medicinal_plants = f'{obj.first_medicinal_plant}, {obj.second_medicinal_plant}.'
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            flower_plants,
            medicinal_plants
        ]
        worksheet.append(row)
    
    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create a response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="radiant_green_corner_data.xlsx"'
    
    return response
    
def access_control_view(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        for key, value in data.items():
            try:
                obj = AccessControl.objects.get(key=key)
                obj.is_active = value
                obj.save()
            except Exception as e:
                messages.error(request, f"Error updating {key}: {str(e)}")
                return redirect('access_control')
        messages.success(request, f"Successfully updated {key}.")
        return redirect('access_control')
                
    if request.method == 'GET':
        access_controls = AccessControl.objects.all()
        access_control_states = {item.key: item.is_active for item in access_controls}
        return render(request, 'access_control.html', {'access_control': access_control_states})


@login_required
def doctors_opinion_view(request):
    if request.method == 'GET':
        search_query = request.GET.get('search', '')
        page_number = int(request.GET.get('page') or 1)
        per_page = int(request.GET.get("per_page") or 10)
        sort = request.GET.get("sort", "territory")
        direction = request.GET.get("direction", "asc")
        
        data = DoctorOpinion.objects.select_related('territory').all()
        try:
            profile = request.user.userprofile
            if profile.user_type == 'zone':
                data = data.filter(territory__zone_name=profile.zone_name)
            elif  profile.user_type == 'region':
                data = data.filter(territory__region_name=profile.region_name)
        except UserProfile.DoesNotExist:
            if not request.user.is_superuser:
                data = DoctorOpinion.objects.none()
                
        if search_query:
            data = data.filter(
                Q(dr_id__icontains=search_query) |
                Q(dr_name__icontains=search_query) |
                Q(territory__territory__icontains=search_query) |
                Q(territory__territory_name__icontains=search_query) |
                Q(territory__region_name__icontains=search_query) |
                Q(territory__zone_name__icontains=search_query) | 
                Q(dr_address__icontains=search_query) |
                Q(dr_phone__icontains=search_query)
            )
        
        sort_by = sort
        if sort_by == "territory":
            sort_by = "territory__territory"
        elif sort_by == "territory_name":
            sort_by = "territory__territory_name"
        elif sort_by == "region":
            sort_by = "territory__region_name"
        elif sort_by == "zone":
            sort_by = "territory__zone_name"
        elif sort_by == "dr_id":
            sort_by = "dr_id"
        elif sort_by == "dr_name":
            sort_by = "dr_name"
        if direction == "desc":
            sort_by = f"-{sort_by}"
            
        data = data.order_by(sort_by)
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)
        context = {
            'data': page_obj,
            'search_query': search_query,
            'sort': sort,
            'direction': direction,
            'per_page': per_page,
            'page_number': page_number,
            'total_pages': paginator.num_pages,
        }
        return render(request, 'doctors_opinion.html', context)

@login_required 
def dop_export(request):
    """
    Export the Doctor's Opinion data to an Excel file.
    """

    # Create a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Doctor's Opinion Data"
    
    # Define the header row
    headers = ['Dr. RPL ID', 'Dr. Name','Dr. Address', 'Dr. Mobile', 'Territory ID', 'Territory Name', 'Region', 'Zone', 'Indications']
    worksheet.append(headers)
    
    # Populate the worksheet with data
    queryset = DoctorOpinion.objects.select_related('territory')
    try:
        profile = request.user.userprofile
        if profile.user_type == 'zone':
            queryset = queryset.filter(territory__zone_name=profile.zone_name)
        elif  profile.user_type == 'region':
            queryset = queryset.filter(territory__region_name=profile.region_name)
    except UserProfile.DoesNotExist:
        if not request.user.is_superuser:
            queryset = DoctorOpinion.objects.none()
    for obj in queryset:
        indications_text = "\n".join(f"{i+1}. {ind.indication_text}" for i, ind in enumerate(obj.indications.all()))
        row = [
            obj.dr_id,
            obj.dr_name,
            obj.dr_address,
            obj.dr_phone,
            obj.territory.territory,
            obj.territory.territory_name,
            obj.territory.region_name,
            obj.territory.zone_name,
            indications_text
        ]
        worksheet.append(row)
        # Apply wrapText only to the last cell (Indications column)
        indications_cell = worksheet.cell(row=worksheet.max_row, column=9)  # 9th column = Indications
        indications_cell.alignment = Alignment(wrap_text=True)
    
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 5

    
    # Save the workbook to a BytesIO object
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Create a response with the Excel file
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="doctors_opinion_data.xlsx"'
    
    return response

@login_required
def doctors_data(request):
    if request.method == 'GET':
        # Get query params with default values
        search_query = request.GET.get('search_query', '')
        sort = request.GET.get('sort', 'dr_id')
        direction = request.GET.get('direction', 'asc')
        per_page = int(request.GET.get('per_page', 10))
        page_number = int(request.GET.get('page_number', 1))

        # Base queryset
        data = Doctor.objects.all()

        # Filtering (search)
        if search_query:
            data = data.filter(
                Q(id__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(speciality__icontains=search_query) |
                Q(designation__icontains=search_query) |
                Q(chambers__district__icontains=search_query) |
                Q(chambers__upazila__icontains=search_query) |
                Q(chambers__thana__icontains=search_query) |
                Q(chambers__address__icontains=search_query) |
                Q(chambers__phone__icontains=search_query)
            ).distinct()

        # Sorting
        sort_by = {
            "dr_id": "id",
            "dr_name": "name"
        }.get(sort, sort)

        if direction == "desc":
            sort_by = f"-{sort_by}"

        data = data.order_by(sort_by)

        # Pagination
        paginator = Paginator(data, per_page)
        page_obj = paginator.get_page(page_number)

        # Serialize chambers to JSON for each doctor
        for doctor in page_obj:
            chambers_list = []
            for ch in doctor.chambers.all():
                chambers_list.append({
                    "address": ch.address,
                    "phone": ch.phone,
                    "division": ch.division,
                    "district": ch.district,
                    "upazila": ch.upazila,
                    "thana": ch.thana,
                })
            doctor.chambers_json = json.dumps(chambers_list)

        # Pass context to template
        context = {
            'data': page_obj,
            'search_query': search_query,
            'sort': sort,
            'direction': direction,
            'per_page': per_page,
            'page_number': page_number,
            'total_pages': paginator.num_pages,
        }

        return render(request, 'doctors.html', context)

@login_required
def doctors_data_export(request):
    wb = openpyxl.Workbook()

    # --- Sheet 1: Doctors ---
    ws_doctors = wb.active
    ws_doctors.title = "Doctors"
    ws_doctors.append(['ID', 'Name', 'Speciality', 'Designation'])

    doctors = Doctor.objects.all()
    for doc in doctors:
        ws_doctors.append([doc.id, doc.name, doc.speciality, doc.designation])

    # --- Sheet 2: Chambers ---
    ws_chambers = wb.create_sheet(title="Chambers")
    ws_chambers.append(['Doctor ID', 'Doctor Name', 'Address', 'Phone', 'Division', 'District', 'Upazila', 'Thana'])

    chambers = Chamber.objects.select_related('doctor').all()
    for ch in chambers:
        ws_chambers.append([
            ch.doctor.id,
            ch.doctor.name,
            ch.address,
            ch.phone,
            ch.division,
            ch.district,
            ch.upazila,
            ch.thana
        ])

    # Set response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=doctors_and_chambers.xlsx'
    wb.save(response)
    return response

       